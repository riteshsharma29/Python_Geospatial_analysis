import pandas as pd
from openpyxl import load_workbook
from opencage.geocoder import OpenCageGeocode
import sys

'''
https://amaral.northwestern.edu/blog/getting-long-lat-list-cities
https://pythoninoffice.com/use-python-to-combine-multiple-excel-files/
'''

# key = "b3b45bc8cb4e4622a58c67b5bba7e8d3"  # get api key from:  https://opencagedata.com
key = "c4956bfaf3184517abd77ffa08102963"
geocoder = OpenCageGeocode(key)

def get_lat_long(cityname):
    results = geocoder.geocode(cityname)
    lat = results[0]['geometry']['lat']
    lng = results[0]['geometry']['lng']
    return lat, lng

# Load excel Workbook using openpyxl
book = load_workbook("US_Source_states_cities.xlsx")
# worksheets = tuple(book.sheetnames)
worksheets = ('Hawaii', 'Idaho', 'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana', 'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota', 'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada', 'New Hampshire', 'New Jersey', 'New Mexico', 'New York', 'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon', 'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota', 'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington', 'West Virginia', 'Wisconsin', 'Wyoming')

for sheetname in worksheets:
    if sheetname != 'Sheet2':
        df = pd.read_excel("US_Source_states_cities.xlsx",sheet_name=sheetname)
        city = tuple(df['city'].to_list())
        l_1 = []
        l_2 = []
        l_3 = []
        for c in city:
            lat_val,long_val = get_lat_long(c + ", United states of America")
            l_1.append(c)
            l_2.append(lat_val)
            l_3.append(long_val)
        tmp_df = pd.DataFrame({"city":l_1,"latitude":l_2,"longitude":l_3})
        tmp_df.to_excel(sheetname + ".xlsx",sheet_name=sheetname,index=False)
        print(sheetname)




