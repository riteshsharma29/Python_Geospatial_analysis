import streamlit as st
from openpyxl import load_workbook
import folium
import streamlit.components.v1 as components
import pandas as pd
import shutil
import os

# Load excel Workbook using openpyxl
book = load_workbook("all_US_states.xlsx")
worksheets = book.sheetnames

# layout
st.title("Streamlit Geospatial Analysis app")
st.markdown("**USA STATES and its respective cities**")
choice = st.sidebar.radio('SELECT US STATE',worksheets)

def create_map(state):
    cities = pd.read_excel("all_US_states.xlsx", sheet_name=state)
    world_all_cities_tooltip = folium.Map(
        zoom_start=2,
        location=[13.133932434766733, 16.103938729508073]
    )
    for _, city in cities.iterrows():
        folium.Marker(
            location=[city['latitude'], city['longitude']],
            popup=city['city'],
            tooltip=city['city'],
        ).add_to(world_all_cities_tooltip)
    world_all_cities_tooltip.save("C:\\xampp\\htdocs\\" + state + ".html")
    components.iframe("http://localhost/" + state + ".html", width=900, height=500)
    st.markdown(state + " State  : Cities are " + str(cities['city'].to_list()))

create_map(choice)
