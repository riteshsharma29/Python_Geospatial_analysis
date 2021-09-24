import os
import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter
import xlsxwriter

cwd = os.path.abspath('')
files = os.listdir(cwd)

#Create Excel Workbook
workbook = xlsxwriter.Workbook('all_US_states.xlsx')
worksheet = workbook.add_worksheet()
workbook.close()

#Load excel Workbook using openpyxl
book = load_workbook('all_US_states.xlsx')
writer = ExcelWriter('all_US_states.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

df = pd.DataFrame()
for file in files:
    if file.endswith('.xlsx'):
        sheetname = file.replace(".xlsx","")
        df = pd.read_excel(file, sheet_name=sheetname)
        df.to_excel(writer, sheet_name=sheetname, index=False, header=True)

first_sheet = book['Sheet1']
book.remove(first_sheet)
writer.save()
